import pandas as pd
from pulp import *
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import os

# #ror data: rooms, time slots, days
lecture_rooms = ["R100A", "R100B", "R100C", "R100D", "R100E", "R100F"]
lab_rooms = ["Lab1", "Lab2", "Lab3", "Lab4", "Lab5", "Lab6", "Hyflex1", "Hyflex2"]
timeslots = list(range(1, 46))  # 45 slots: 5 days × 9 hours
lunch_slots = [5, 14, 23, 32, 41]
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
times = ["8-9","9-10","10-11","11-12","12-1","1-2","2-3","3-4","4-5"]

# #ror scheduling patterns: 3-unit (MWF, MTTh, etc) & 2-unit (MW, TTh, WF)
scheduling_patterns = {
    3: [[1,19,37], [2,20,38], [3,21,39], [4,22,40], [5,23,41], [6,24,42], [7,25,43], [8,26,44], [9,27,45], #MWF
        [1,10,19], [2,11,20], [3,12,21], [4,13,22], [5,14,23], [6,15,24], [7,16,25], [8,17,26], [9,18,27], #MTTh
        [10,19,28], [11,20,29], [12,21,30], [13,22,31], [14,23,32], [15,24,33], [16,25,34], [17,26,35], [18,27,36], #TTTh
        [19,28,37], [20,29,38], [21,30,39], [22,31,40], [23,32,41], [24,33,42], [25,34,43], [26,35,44], [27,36,45] #WThF
    ],
    2: [[1,19], [2,20], [3,21], [4,22], [5,23], [6,24], [7,25], [8,26], [9,27], #MW
        [10,28], [11,29], [12,30], [13,31], [14,32], [15,33], [16,34], [17,35], [18,36], #TTh
        [19,37], [20,38], [21,39], [22,40], [23,41], [24,42], [25,43], [26,44], [27,45] #WF
    ]
}
# #ror lab slots: 3 consecutive hours, can start at these times
valid_lab_starts = [1,4,7,10,13,16,19,22,25,28,31,34,37,40,43]
# #ror all sections across CS & IT programs
sections = [
    "CS1A", "CS1B",
    "CS2A", "CS2B",
    "CS3A", "CS3B",
    "CS4A", "CS4B",
    "IT1A", "IT1B", "IT1C",
    "IT2A", "IT2B", "IT2C",
    "IT3A", "IT3B", "IT3C",
    "IT4A", "IT4B", "IT4C"
]

# #ror lecture subjects with unit counts
lecture_subjects = {
    "Computer Programming 1": 2,
    "Computer Science Fundamentals": 2,
    "Data Structures": 2,
    "Digital Design": 2,
    "Database Systems": 2,
    "Discrete Structures": 2,
    "Operating Systems": 2,
    "System Analysis and Design": 2,
    "Computer Networks": 2,
    "Artificial Intelligence": 2,
    "Programming Languages": 2,
    "Modelling and Simulation": 3,
    "Special Topics": 3,
    "Thesis Writing I": 3,
    "Human Computer Interaction": 2,

    "Introduction to Computing": 2,
    "Science, Technology, and Society": 3,
    "Understanding the Self": 3,          # kept once (was duplicated)
    "Reading Visual Art": 3,
    "Movement Competency Training (MCT)": 2,
    "CWTS 1/ROTC 1": 3,
    "Database Systems": 2,
    "Operating System": 2,
    "Data Structures & Algorithms": 2,
    "Introduction to Game Development": 2,
    "Ethics": 3,
    "Fundamentals of Accounting for IT": 3,
    "Environmental Science": 3,
    "Dance": 2,
    "Computer Networks 1": 2,
    "Platform Technologies": 2,
    "Data Analytics": 3,
    "Information and Project Management": 2,
    "System Administration & Maintenance": 3,
    "Fundamentals of Business Analytics": 3,
    "Life and Works of Rizal": 3,
    "Social Issues & Ethics in Computing": 3,
    "Information Assurance & Security 2": 2,
    "Multimedia Systems": 2,
    "IT Seminar": 2,
    "Capstone Project Writing": 2,
}

# #ror lab subjects (1 unit each)
lab_subjects = {
    "Computer Programming 1 Lab": 1,
    "Computer Science Fundamentals Lab": 1,
    "Data Structures Lab": 1,
    "Digital Design Lab": 1,
    "Database Systems Lab": 1,
    "Discrete Structures Lab": 1,
    "Operating System Lab": 1,
    "System Analysis and Design Lab": 1,
    "Computer Networks Lab": 1,
    "Artificial Intelligence Lab": 1,
    "Programming Languages Lab": 1,
    "Human Computer Interaction Lab": 1,
    "Compiler Design Lab": 1,
    # IT subjects
    "Introduction to Computing Lab": 1,
    "Introduction to Game Development Lab": 1,
    "Computer Networks 1 Lab": 1,
    "Platform Technologies Lab": 1,
    "Data Analytics Lab": 1,
    "Information and Project Management Lab": 1,
    "System Administration & Maintenance Lab": 1,
    "Fundamentals of Business Analytics Lab": 1,
    "Information Assurance & Security 2 Lab": 1,
    "Multimedia Systems Lab": 1,
    "IT Seminar Lab": 1,
    "Capstone Project Writing Lab": 1,
}

# #ror curriculum subjects per section (from official curriculum)
section_subjects = {
    "CS1A": [
        "Computer Science Fundamentals", "Computer Programming 1",
        "Science, Technology, and Society", "Understanding the Self",
        "Reading Visual Art",
        "Computer Science Fundamentals Lab",
        "Computer Programming 1 Lab",
    ],
    "CS1B": [
        "Computer Science Fundamentals", "Computer Programming 1",
        "Science, Technology, and Society", "Understanding the Self",
        "Reading Visual Art",
        "Computer Science Fundamentals Lab",
        "Computer Programming 1 Lab",
    ],
    "CS2A": [
        "Data Structures", "Digital Design", "Database Systems",
        "Discrete Structures", "Ethics", "Environmental Science",

        "Data Structures Lab", "Digital Design Lab",
        "Database Systems Lab", "Discrete Structures Lab",
    ],
    "CS2B": [
        "Data Structures", "Digital Design", "Database Systems",
        "Discrete Structures", "Ethics", "Environmental Science",

        "Data Structures Lab", "Digital Design Lab",
        "Database Systems Lab", "Discrete Structures Lab",
    ],
    "CS3A": [
        "Operating Systems", "System Analysis and Design",
        "Computer Networks", "Artificial Intelligence",
        "Life and Works of Rizal",
        "Operating System Lab", "System Analysis and Design Lab",
        "Computer Networks Lab", "Artificial Intelligence Lab",
    ],
    "CS3B": [
        "Operating Systems", "System Analysis and Design",
        "Computer Networks", "Artificial Intelligence",
        "Life and Works of Rizal",
        "Operating System Lab", "System Analysis and Design Lab",
        "Computer Networks Lab", "Artificial Intelligence Lab",
    ],
    "CS4A": [
        "Programming Languages", "Human Computer Interaction",
        "Special Topics", "Thesis Writing I",

        "Programming Languages Lab",
        "Human Computer Interaction Lab",
    ],
    "CS4B": [
        "Programming Languages", "Human Computer Interaction",
        "Special Topics", "Thesis Writing I",

        "Programming Languages Lab",
        "Human Computer Interaction Lab",
    ],
    "IT1A": [
        "Introduction to Computing", "Computer Programming 1",
        "Science, Technology, and Society", "Understanding the Self",
        "Reading Visual Art",
        "Platform Technologies Lab", "Computer Programming 1 Lab",
    ],
    "IT1B": [
        "Introduction to Computing", "Computer Programming 1",
        "Science, Technology, and Society", "Understanding the Self",
        "Reading Visual Art",
        "Platform Technologies Lab", "Computer Programming 1 Lab",
    ],
    "IT1C": [
        "Introduction to Computing", "Computer Programming 1",
        "Science, Technology, and Society", "Understanding the Self",
        "Reading Visual Art",
        "Platform Technologies Lab", "Computer Networks 1 Lab",
    ],
    "IT2A": [
        "Database Systems", "Operating System", "Data Structures & Algorithms",
        "Introduction to Game Development", "Ethics",
        "Fundamentals of Accounting for IT", "Environmental Science",
        "Database Systems Lab", "Operating System Lab",
        "Data Structures Lab", "Introduction to Game Development Lab",
    ],
    "IT2B": [
        "Database Systems", "Operating System", "Data Structures & Algorithms",
        "Introduction to Game Development", "Ethics",
        "Fundamentals of Accounting for IT", "Environmental Science",
        "Database Systems Lab", "Operating System Lab",
        "Data Structures Lab", "Introduction to Game Development Lab",
    ],
    "IT2C": [
        "Database Systems", "Operating System", "Data Structures & Algorithms",
        "Introduction to Game Development", "Ethics",
        "Fundamentals of Accounting for IT", "Environmental Science",
        "Database Systems Lab", "Operating System Lab",
        "Data Structures Lab", "Introduction to Game Development Lab",
    ],
    "IT3A": [
        "Computer Networks 1", "Platform Technologies", "Data Analytics",
        "Information and Project Management", "System Administration & Maintenance",
        "Fundamentals of Business Analytics", "Life and Works of Rizal",
        "Computer Networks 1 Lab", "Platform Technologies Lab", "Information and Project Management Lab", 
    ],
    "IT3B": [
        "Computer Networks 1", "Platform Technologies", "Data Analytics",
        "Information and Project Management", "System Administration & Maintenance",
        "Fundamentals of Business Analytics", "Life and Works of Rizal",
        "Computer Networks 1 Lab", "Platform Technologies Lab", "Information and Project Management Lab",
    ],
    "IT3C": [
        "Computer Networks 1", "Platform Technologies", "Data Analytics",
        "Information and Project Management", "System Administration & Maintenance",
        "Fundamentals of Business Analytics", "Life and Works of Rizal",
        "Computer Networks 1 Lab", "Platform Technologies Lab", "Information and Project Management Lab",
    ],
    "IT4A": [
        "Social Issues & Ethics in Computing", "Information Assurance & Security 2",
        "Multimedia Systems", "IT Seminar", "Capstone Project Writing",
        "Information Assurance & Security 2 Lab", "Multimedia Systems Lab",
        "IT Seminar Lab", "Capstone Project Writing Lab",
    ],
    "IT4B": [
        "Social Issues & Ethics in Computing", "Information Assurance & Security 2",
        "Multimedia Systems", "IT Seminar", "Capstone Project Writing",
        "Information Assurance & Security 2 Lab", "Multimedia Systems Lab",
        "IT Seminar Lab", "Capstone Project Writing Lab",
    ],
    "IT4C": [
        "Social Issues & Ethics in Computing", "Information Assurance & Security 2",
        "Multimedia Systems", "IT Seminar", "Capstone Project Writing",
        "Information Assurance & Security 2 Lab", "Multimedia Systems Lab",
        "IT Seminar Lab", "Capstone Project Writing Lab",
    ],
}

# #ror separate lectures from labs by section
lecture_subjects_in_section = {
    sec: [sub for sub in section_subjects.get(sec, []) if sub in lecture_subjects]
    for sec in sections
}
lab_subjects_in_section = {
    sec: [sub for sub in section_subjects.get(sec, []) if sub in lab_subjects]
    for sec in sections
}

# #ror helper functions for slot conversion
def get_day_from_slot(slot):
    return (slot - 1) // 9

def get_time_from_slot(slot):
    return (slot - 1) % 9

#Model start
print("Room Allocation Optimization for the First Semester 2026 - 2027 for CCIS Programs")
print("=" * 60)

prob = LpProblem("RoomAllocation", LpMinimize)

# #ror decision variables: lecture assignments (section, subject, room, pattern)
x_lecture = {}
for section in sections:
    for subject in lecture_subjects_in_section[section]:
        units = lecture_subjects[subject]
        patterns = scheduling_patterns.get(units, [])
        for pattern_idx, pattern in enumerate(patterns):
            for room in lecture_rooms:
                var_name = f"x_lec_{section}_{subject}_{room}_{pattern_idx}"
                x_lecture[(section, subject, room, pattern_idx)] = LpVariable(var_name, cat='Binary')

# #ror decision variables: lab assignments (section, subject, room, start_slot)
y_lab = {}
for section in sections:
    for subject in lab_subjects_in_section[section]:
        for room in lab_rooms:
            for start_slot in valid_lab_starts:
                three_slots = [start_slot, start_slot + 1, start_slot + 2]
                if all(slot < 46 for slot in three_slots):
                    var_name = f"y_lab_{section}_{subject}_{room}_{start_slot}"
                    y_lab[(section, subject, room, start_slot)] = LpVariable(var_name, cat='Binary')

# #ror binary vars: track if section occupies a slot
occupy = {}
for sec in sections:
    for slot in timeslots:
        occupy[(sec, slot)] = LpVariable(f"occ_{sec}_{slot}", cat="Binary")
        
print("\nConstraints.")
constraint_count = 0

# #ror constraint 1: no overlapping classes per section

for sec in sections:
    for slot in timeslots:

        involved_vars = []

        #for each subject in the section, check if it can be in this slot based on its patterns (for lectures) or 3-slot window (for labs)
        for subject in lecture_subjects_in_section[sec]:
            units = lecture_subjects[subject]
            patterns = scheduling_patterns.get(units, [])

            for p_idx, pattern in enumerate(patterns):
                for room in lecture_rooms:
                    key = (sec, subject, room, p_idx)
                    x_var = x_lecture.get(key)

                    if x_var is not None and slot in pattern:
                        involved_vars.append(x_var)

        for subject in lab_subjects_in_section[sec]:
            for room in lab_rooms:
                for start_slot in valid_lab_starts:
                    key = (sec, subject, room, start_slot)
                    y_var = y_lab.get(key)

                    if y_var is not None:
                        if slot in [start_slot, start_slot + 1, start_slot + 2]:
                            involved_vars.append(y_var)

        if involved_vars:
            prob += lpSum(involved_vars) <= 1, f"slot_exclusive_{sec}_{slot}"
            constraint_count += 1

# #ror penalty vars: track dense class windows (penalize in objective)
overload = {}
for sec in sections:
    for day in range(5):
        for start in range(4):  # 6-hour windows
            overload[(sec, day, start)] = LpVariable(
                f"overload_{sec}_{day}_{start}",
                cat="Binary"
            )

# #ror constraint 2: every lecture subject assigned exactly once
print("2. Every subject must be assigned")
for section in sections:
    for subject in lecture_subjects_in_section[section]:
        units = lecture_subjects[subject]
        patterns = scheduling_patterns.get(units, [])
        
        #Sum of all (room, pattern) assignments for this subject = 1
        subject_vars = [x_lecture[(section, subject, room, p_idx)]
                       for p_idx in range(len(patterns))
                       for room in lecture_rooms
                       if (section, subject, room, p_idx) in x_lecture]
        
        if subject_vars:
            prob += lpSum(subject_vars) == 1, f"lec_assign_{section}_{subject}"
            constraint_count += 1

#every lab subhect must be assigned, and assigned at one of the lab starts
print("2. Each lab subject exactly once")
for section in sections:
    for subject in lab_subjects_in_section[section]:
        subject_vars = [y_lab[(section, subject, room, start_slot)]
                       for room in lab_rooms
                       for start_slot in valid_lab_starts
                       if (section, subject, room, start_slot) in y_lab]
        
        if subject_vars:
            prob += lpSum(subject_vars) == 1, f"lab_assign_{section}_{subject}"
            constraint_count += 1

# #ror constraint 3: no lecture room double-booking
print("3. No lecture room double-booking")
for room in lecture_rooms:
    for slot in timeslots:
        if slot in lunch_slots:
            continue
        
        #find all lecture assignments that use this room at this slot
        room_slot_vars = []
        for section in sections:
            for subject in lecture_subjects_in_section[section]:
                units = lecture_subjects[subject]
                patterns = scheduling_patterns.get(units, [])
                
                for p_idx, pattern in enumerate(patterns):
                    if slot in pattern:
                        key = (section, subject, room, p_idx)
                        if key in x_lecture:
                            room_slot_vars.append(x_lecture[key])
        
        if room_slot_vars:
            prob += lpSum(room_slot_vars) <= 1, f"lec_room_slot_{room}_{slot}"
            constraint_count += 1

# #ror constraint 4: no lab room double-booking
print("4. No lab room double-booking")
for room in lab_rooms:
    for slot in timeslots:
        # Find all lab assignments that use this room at this slot
        room_slot_vars = []
        for section in sections:
            for subject in lab_subjects_in_section[section]:
                for start_slot in valid_lab_starts:
                    three_slots = [start_slot, start_slot + 1, start_slot + 2]
                    if slot in three_slots:
                        key = (section, subject, room, start_slot)
                        if key in y_lab:
                            room_slot_vars.append(y_lab[key])
        
        if room_slot_vars:
            prob += lpSum(room_slot_vars) <= 1, f"lab_room_slot_{room}_{slot}"
            constraint_count += 1

print("5. Adding workload penalty constraints")

for sec in sections:
    for day in range(5):
        base = day * 9

        for start in range(0, 4):  #windows of 6 slots
            window_slots = [base + start + i for i in range(6)]

            involved = []

            for slot in window_slots:
                #LECTURES
                for subject in lecture_subjects_in_section[sec]:
                    units = lecture_subjects[subject]
                    patterns = scheduling_patterns.get(units, [])

                    for p_idx, pattern in enumerate(patterns):
                        for room in lecture_rooms:
                            key = (sec, subject, room, p_idx)
                            x_var = x_lecture.get(key)

                            if x_var is not None and slot in pattern:
                                involved.append(x_var)

                #LABS
                for subject in lab_subjects_in_section[sec]:
                    for room in lab_rooms:
                        for start_slot in valid_lab_starts:
                            key = (sec, subject, room, start_slot)
                            y_var = y_lab.get(key)

                            if y_var is not None:
                                if slot in [start_slot, start_slot + 1, start_slot + 2]:
                                    involved.append(y_var)

            if involved:
                prob += overload[(sec, day, start)] <= lpSum(involved)

print(f"Total constraints: {constraint_count}")

# (penalty 1: room changes between consecutive classes in different rooms)
print("6. Adding room consistency penalty")

room_used = {}
room_change_penalties = []

# Penalize using many different rooms in the same section
for sec in sections:

    all_rooms = lecture_rooms + lab_rooms
    section_subjects_all = section_subjects.get(sec, [])

    for room in all_rooms:

        room_used[(sec, room)] = LpVariable(
            f"room_used_{sec}_{room}",
            cat="Binary"
        )

        room_assignments = []

        # lecture assignments in this room
        for subject in lecture_subjects_in_section[sec]:
            units = lecture_subjects[subject]
            patterns = scheduling_patterns.get(units, [])

            for p_idx in range(len(patterns)):
                key = (sec, subject, room, p_idx)

                if key in x_lecture:
                    room_assignments.append(x_lecture[key])

        # lab assignments in this room
        for subject in lab_subjects_in_section[sec]:
            for start_slot in valid_lab_starts:
                key = (sec, subject, room, start_slot)

                if key in y_lab:
                    room_assignments.append(y_lab[key])

        if room_assignments:
            # if any subject uses this room, activate room_used
            for var in room_assignments:
                prob += room_used[(sec, room)] >= var

            room_change_penalties.append(room_used[(sec, room)])

# total rooms used by a section = penalty
penalty_room_change = lpSum(room_change_penalties)

# (penalty: multiple labs in same day for the same section)
# FAST LAB SAME-DAY PENALTY
print("7. Adding lab clustering penalty")

lab_day_penalties = []

for sec in sections:
    for day in range(5):

        labs_today = []

        for subject in lab_subjects_in_section[sec]:
            for room in lab_rooms:
                for start_slot in valid_lab_starts:

                    key = (sec, subject, room, start_slot)

                    if key in y_lab:
                        if get_day_from_slot(start_slot) == day:
                            labs_today.append(y_lab[key])

        if labs_today:
            # penalty only if more than 1 lab on this day
            overload_lab = LpVariable(
                f"lab_overload_{sec}_{day}",
                lowBound=0,
                cat="Integer"
            )

            # if 1 lab = 0 penalty
            # if 2 labs = 1 penalty
            # if 3 labs = 2 penalty
            prob += overload_lab >= lpSum(labs_today) - 1

            lab_day_penalties.append(overload_lab)

penalty_lab_same_day = lpSum(lab_day_penalties)

print("\nOBJ FUNCTION: Minimize class density, room changes, and lab spread")
penalty_consecutive = lpSum(overload.values())
penalty_combined = penalty_consecutive + penalty_room_change + penalty_lab_same_day

prob += penalty_combined, "Objective"

#solve
print("\nSolving...")
prob.solve(PULP_CBC_CMD(msg=0, timeLimit=300))

print(f"Status: {LpStatus[prob.status]}")

if LpStatus[prob.status] != 'Optimal' and LpStatus[prob.status] != 'Feasible':
    print("ERROR: Could not find a feasible schedule!")
    exit(1)

# extract solution and build schedule data
print("\nExtracting solution...")
schedule_data = []

#extract all lecture assignments
for (section, subject, room, p_idx), var in x_lecture.items():
    if var.varValue == 1:
        units = lecture_subjects[subject]
        patterns = scheduling_patterns.get(units, [])
        pattern = patterns[p_idx]
        
        for slot in pattern:
            day = get_day_from_slot(slot)
            time_idx = get_time_from_slot(slot)
            schedule_data.append({
                'Section': section,
                'Subject': subject,
                'Room': room,
                'Type': 'Lecture',
                'Day': days[day],
                'Time': times[time_idx],
                'Slot': slot,
                'DayNum': day,
                'TimeNum': time_idx
            })

# #ror extract all lab assignments
for (section, subject, room, start_slot), var in y_lab.items():
    if var.varValue == 1:
        three_slots = [start_slot, start_slot + 1, start_slot + 2]
        day = get_day_from_slot(start_slot)
        
        for slot in three_slots:
            time_idx = get_time_from_slot(slot)
            schedule_data.append({
                'Section': section,
                'Subject': subject,
                'Room': room,
                'Type': 'Lab',
                'Day': days[day],
                'Time': times[time_idx],
                'Slot': slot,
                'DayNum': day,
                'TimeNum': time_idx
            })

# #ror build Excel output
print("\nGenerating Excel timetable...")
output_file = r"C:\Operations Research\Schedule_Output.xlsx"

# Remove existing file if it exists (to avoid permission errors)
if os.path.exists(output_file):
    try:
        os.remove(output_file)
        print("Removed existing file...")
    except Exception as e:
        print(f"Warning: Could not remove existing file: {e}")

df = pd.DataFrame(schedule_data)
df_sorted = df.sort_values(['Section', 'DayNum', 'TimeNum']).drop(columns=['DayNum', 'TimeNum'])

# #ror setup Excel workbook & styling
print("Creating workbook...")
wb = Workbook()
wb.remove(wb.active)

# #ror colors & formatting
lecture_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
lab_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# #ror sheet 1: master schedule
print("  Sheet 1: Master Schedule")
ws_master = wb.create_sheet('Master Schedule', 0)
headers = ['Section', 'Subject', 'Room', 'Type', 'Day', 'Time', 'Slot']
for col_idx, header in enumerate(headers, start=1):
    cell = ws_master.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row in enumerate(df_sorted.itertuples(index=False), start=2):
    is_lab = row.Type == 'Lab'
    color = lab_fill if is_lab else lecture_fill
    
    for col_idx, value in enumerate(row, start=1):
        cell = ws_master.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = color
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# #ror auto-fit column widths
for col_idx, header in enumerate(headers, start=1):
    ws_master.column_dimensions[get_column_letter(col_idx)].width = 20

# #ror sheet 2: 9x5 timetable grid
print("  Sheet 2: Timetable")
ws_timetable = wb.create_sheet('Timetable', 1)

# Set column widths
ws_timetable.column_dimensions['A'].width = 12
for col_idx in range(2, 7):
    ws_timetable.column_dimensions[get_column_letter(col_idx)].width = 25

# #ror header row with days
ws_timetable['A1'] = 'TIME'
header_cell = ws_timetable['A1']
header_cell.fill = header_fill
header_cell.font = header_font
header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
header_cell.border = thin_border

for day_idx, day in enumerate(days):
    col_letter = get_column_letter(day_idx + 2)
    cell = ws_timetable[f'{col_letter}1']
    cell.value = day
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws_timetable.row_dimensions[1].height = 25

# #ror fill time slots & classes
for time_idx, time_slot in enumerate(times):
    row_num = time_idx + 2
    ws_timetable.row_dimensions[row_num].height = 60

    # Time column
    time_cell = ws_timetable[f'A{row_num}']
    time_cell.value = time_slot
    time_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    time_cell.border = thin_border
    time_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    time_cell.font = Font(bold=True)

    # #ror for each day, populate cells
    for day_idx in range(5):
        col_letter = get_column_letter(day_idx + 2)
        slot = day_idx * 9 + time_idx + 1

        # Find all classes at this slot
        classes_at_slot = df_sorted[df_sorted['Slot'] == slot]

        cell = ws_timetable[f'{col_letter}{row_num}']
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        if len(classes_at_slot) > 0:
            first_type = classes_at_slot.iloc[0]['Type']
            is_lab = first_type == 'Lab'
            cell.fill = lab_fill if is_lab else lecture_fill

            cell_text = ""
            for idx, row in classes_at_slot.iterrows():
                subject = row['Subject']
                if len(subject) > 25:
                    subject = subject[:22] + "..."
                cell_text += f"{row['Section']}\n{row['Room']}\n{subject}\n"

            cell.value = cell_text.strip()
            cell.font = Font(size=9)

# #ror sheet 3: per section timetables
print("  Creating section sheets...")
for section in sorted(sections):
    section_df = df_sorted[df_sorted['Section'] == section].copy()
    if len(section_df) > 0:
        section_df_sorted = section_df.drop(columns=['Section']).sort_values(['Day', 'Time'])
        
        # Create section list sheet
        sheet_name = f'Sec {section} List'[:31]
        ws_section = wb.create_sheet(sheet_name)
        
        # Headers
        headers = section_df_sorted.columns.tolist()
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_section.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        # Data
        for row_idx, row in enumerate(section_df_sorted.itertuples(index=False), start=2):
            is_lab = row.Type == 'Lab'
            color = lab_fill if is_lab else lecture_fill
            
            for col_idx, value in enumerate(row, start=1):
                cell = ws_section.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = color
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # #ror auto-fit columns
        for col_idx in range(1, len(headers) + 1):
            ws_section.column_dimensions[get_column_letter(col_idx)].width = 18
    
    # #ror per-section 9x5 grid timetable
    grid_sheet_name = f'{section} Timetable'[:31]
    ws_grid = wb.create_sheet(grid_sheet_name)
    
    # Set up the grid
    ws_grid.column_dimensions['A'].width = 12
    for col_idx in range(2, 7):
        ws_grid.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # #ror header row with day names
    ws_grid['A1'] = 'TIME'
    header_cell = ws_grid['A1']
    header_cell.fill = header_fill
    header_cell.font = header_font
    header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_cell.border = thin_border
    
    for day_idx, day in enumerate(days):
        col_letter = get_column_letter(day_idx + 2)
        cell = ws_grid[f'{col_letter}1']
        cell.value = day
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws_grid.row_dimensions[1].height = 25
    
    # #ror fill time slots with classes
    section_schedule = df_sorted[df_sorted['Section'] == section].copy()
    
    for time_idx, time_slot in enumerate(times):
        row_num = time_idx + 2
        ws_grid.row_dimensions[row_num].height = 50
        
        # #ror time label column
        time_cell = ws_grid[f'A{row_num}']
        time_cell.value = time_slot
        time_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        time_cell.border = thin_border
        time_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        time_cell.font = Font(bold=True, size=10)
        
        # For each day
        for day_idx in range(5):
            col_letter = get_column_letter(day_idx + 2)
            slot = day_idx * 9 + time_idx + 1
            
            # #ror fetch classes at this slot
            classes_at_slot = section_schedule[section_schedule['Slot'] == slot]
            
            cell = ws_grid[f'{col_letter}{row_num}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if len(classes_at_slot) > 0:
                # Get subject and room
                first_row = classes_at_slot.iloc[0]
                subject = first_row['Subject']
                room = first_row['Room']
                is_lab = first_row['Type'] == 'Lab'
                
                # Shorten subject name if too long
                if len(subject) > 20:
                    subject = subject[:17] + "..."
                
                cell.value = f"{subject}\n{room}"
                cell.font = Font(size=9, bold=True)
                cell.fill = lab_fill if is_lab else lecture_fill
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# #ror sheet 4: room utilization summary
print("  Sheet 4: Room Utilization")
ws_rooms = wb.create_sheet('Room Utilization')
room_util_data = []
for room in lecture_rooms + lab_rooms:
    room_df = df[df['Room'] == room]
    room_util_data.append({
        'Room': room,
        'Type': 'Lecture' if room in lecture_rooms else 'Lab',
        'Classes': len(room_df),
        'Sections': room_df['Section'].nunique(),
        'Subjects': room_df['Subject'].nunique()
    })

room_util_df = pd.DataFrame(room_util_data)
headers = room_util_df.columns.tolist()
for col_idx, header in enumerate(headers, start=1):
    cell = ws_rooms.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row in enumerate(room_util_df.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws_rooms.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

for col_idx in range(1, len(headers) + 1):
    ws_rooms.column_dimensions[get_column_letter(col_idx)].width = 15

# #ror sheet 5: subject assignment coverage
print("  Sheet 5: Subject Coverage")
ws_coverage = wb.create_sheet('Subject Coverage')
subject_coverage = []
for section in sorted(sections):
    for subject in section_subjects[section]:
        assigned = len(df[(df['Section'] == section) & (df['Subject'] == subject)]) > 0
        subject_coverage.append({
            'Section': section,
            'Subject': subject,
            'Status': '✓ Scheduled' if assigned else '✗ NOT Scheduled'
        })

coverage_df = pd.DataFrame(subject_coverage)
headers = coverage_df.columns.tolist()
for col_idx, header in enumerate(headers, start=1):
    cell = ws_coverage.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row in enumerate(coverage_df.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws_coverage.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

for col_idx in range(1, len(headers) + 1):
    ws_coverage.column_dimensions[get_column_letter(col_idx)].width = 25

# #ror save Excel file
try:
    wb.save(output_file)
    print(f"✓ Excel file saved: {output_file}")
except Exception as e:
    print(f"✗ Error saving Excel file: {e}")
    print("  Trying alternative location...")
    alt_file = r"C:\Operations Research\Schedule_Output_alt.xlsx"
    try:
        wb.save(alt_file)
        print(f"Excel file saved to: {alt_file}")
        output_file = alt_file
    except Exception as e2:
        print(f"Could not save to alternative location either: {e2}")

# Print summary
print("\n" + "=" * 60)
print("SCHEDULE SUMMARY")
print("=" * 60)
for section in sorted(sections):
    section_classes = len(df[df['Section'] == section])
    print(f"{section}: {section_classes} class sessions")

total_sessions = len(df)
print(f"\nTotal class sessions: {total_sessions}")
print(f"Total subjects assigned: {df['Subject'].nunique()}")
print(f"Rooms utilized: {df['Room'].nunique()}")
