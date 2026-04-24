import pandas as pd
from pulp import *
from collections import defaultdict
from openpyxl.styles import PatternFill

# ----------------------------
# DATA
# ----------------------------
lecture_rooms = ["R100A", "R100B", "R100C", "R100D", "R100E", "R100F"]
lab_rooms = ["Lab1", "Lab2", "Lab3", "Lab4", "Lab5", "Lab6", "Hyflex1", "Hyflex2"]
timeslots = list(range(1, 46))  # 5 days x 9
lunch_slots = [5, 14, 23, 32, 41]
sections = [
    "CS1A", "CS1B",
    "CS2A", "CS2B"
]

# BUG FIX 1: lecture_subjects had duplicate key "Understanding the Self"
# (second entry silently overwrote the first in a plain dict).
# Use an OrderedDict-style list-of-tuples then convert so duplicates are explicit.
# Also removed "Understanding the Self" duplicate and kept one copy.
lecture_subjects = {
    # CS subjects
    "Computer Programming 1": 2,
    "Computer Science Fundamentals": 2,
    "Data Structures": 2,
    "Digital Design": 2,
    "Database Systems": 2,
    "Discrete Structures": 2,
    "Operating Systems": 2,
    "System Analysis and Design": 2,
    "Computer Networks": 2,
    "Artificial Intelligence": 2,
    "Programming Languages": 2,
    "Modelling and Simulation": 3,
    "Special Topics": 3,
    "Thesis Writing I": 3,
    "Human Computer Interaction": 2,
    # IT subjects
    "Introduction to Computing": 2,
    "Science, Technology, and Society": 3,
    "Understanding the Self": 3,          # kept once (was duplicated)
    "Reading Visual Art": 3,
    "Movement Competency Training (MCT)": 2,
    "CWTS 1/ROTC 1": 3,
    "Database Systems": 2,
    "Operating System": 2,
    "Data Structures & Algorithms": 2,
    "Introduction to Game Development": 2,
    "Ethics": 3,
    "Fundamentals of Accounting for IT": 3,
    "Environmental Science": 3,
    "Dance": 2,
    "Computer Networks 1": 2,
    "Platform Technologies": 2,
    "Data Analytics": 2,
    "Information and Project Management": 2,
    "System Administration & Maintenance": 2,
    "Fundamentals of Business Analytics": 2,
    "Life and Works of Rizal": 3,
    "Social Issues & Ethics in Computing": 3,
    "Information Assurance & Security 2": 2,
    "Multimedia Systems": 2,
    "IT Seminar": 2,
    "Capstone Project Writing": 2,
}

lab_subjects = {
    "Computer Programming 1 Lab": 1,
    "Computer Science Fundamentals Lab": 1,
    "Data Structures Lab": 1,
    "Digital Design Lab": 1,
    "Database Systems Lab": 1,
    "Discrete Structures Lab": 1,
    "Operating System Lab": 1,
    "System Analysis and Design Lab": 1,
    "Computer Networks Lab": 1,
    "Artificial Intelligence Lab": 1,
    "Programming Languages Lab": 1,
    "Human Computer Interaction Lab": 1,
    "Compiler Design Lab": 1,
    # IT subjects
    "Introduction to Computing Lab": 1,
    "Introduction to Game Development Lab": 1,
    "Computer Networks 1 Lab": 1,
    "Platform Technologies Lab": 1,
    "Data Analytics Lab": 1,
    "Information and Project Management Lab": 1,
    "System Administration & Maintenance Lab": 1,
    "Fundamentals of Business Analytics Lab": 1,
    "Information Assurance & Security 2 Lab": 1,
    "Multimedia Systems Lab": 1,
    "IT Seminar Lab": 1,
    "Capstone Project Writing Lab": 1,
}

section_subjects = {
    "CS1A": [
        "Computer Science Fundamentals", "Computer Programming 1",
        "Science, Technology, and Society", "Understanding the Self",
        "Intensive English", "Reading Visual Arts",

        "Computer Science Fundamentals Lab",
        "Computer Programming 1 Lab",
    ],
    "CS1B": [
        "Computer Science Fundamentals", "Computer Programming 1",
        "Science, Technology, and Society", "Understanding the Self",
        "Intensive English", "Reading Visual Arts",

        "Computer Science Fundamentals Lab",
        "Computer Programming 1 Lab",
    ],
    "CS2A": [
        "Data Structures", "Digital Design", "Database Systems",
        "Discrete Structures", "Ethics", "Environmental Science",

        "Data Structures Lab", "Digital Design Lab",
        "Database Systems Lab", "Discrete Structures Lab",
    ],
    "CS2B": [
        "Data Structures", "Digital Design", "Database Systems",
        "Discrete Structures", "Ethics", "Environmental Science",

        "Data Structures Lab", "Digital Design Lab",
        "Database Systems Lab", "Discrete Structures Lab",
    ],
    "CS3A": [
        "Operating Systems", "Systems Analysis and Design",
        "Computer Networks", "Artificial Intelligence",
        "Life and Works of Rizal", "The Entrepreneurial Mind",

        "Operating Systems Lab", "Systems Analysis and Design Lab",
        "Computer Networks Lab", "Artificial Intelligence Lab",
    ],
    "CS3B": [
        "Operating Systems", "Systems Analysis and Design",
        "Computer Networks", "Artificial Intelligence",
        "Life and Works of Rizal", "The Entrepreneurial Mind",

        "Operating Systems Lab", "Systems Analysis and Design Lab",
        "Computer Networks Lab", "Artificial Intelligence Lab",
    ],
    "CS4A": [
        "Programming Languages", "Human Computer Interaction",
        "Special Topics", "Thesis Writing I",

        "Programming Languages Lab",
        "Human Computer Interaction Lab",
    ],
    "CS4B": [
        "Programming Languages", "Human Computer Interaction",
        "Special Topics", "Thesis Writing I",

        "Programming Languages Lab",
        "Human Computer Interaction Lab",
    ],
    "IT1A": [
        "Introduction to Computing", "Computer Programming I",
        "Science, Technology, and Society", "Understanding the Self",
        "Reading Visual Art",
        "Platform Technologies Lab", "Computer Programming I Lab",
    ],
    "IT1B": [
        "Introduction to Computing", "Computer Programming I",
        "Science, Technology, and Society", "Understanding the Self",
        "Reading Visual Art",
        "Platform Technologies Lab", "Computer Programming I Lab",
    ],
    "IT1C": [
        "Introduction to Computing", "Computer Programming I",
        "Science, Technology, and Society", "Understanding the Self",
        "Reading Visual Art",
        "Platform Technologies Lab", "Computer Networks I Lab",
    ],
    "IT2A": [
        "Database Systems", "Operating System", "Data Structures & Algorithms",
        "Introduction to Game Development", "Ethics",
        "Fundamentals of Accounting for IT", "Environmental Science",
        "Database Systems Lab", "Operating System Lab",
        "Data Structures Lab", "Introduction to Game Development Lab",
    ],
    "IT2B": [
        "Database Systems", "Operating System", "Data Structures & Algorithms",
        "Introduction to Game Development", "Ethics",
        "Fundamentals of Accounting for IT", "Environmental Science",
        "Database Systems Lab", "Operating System Lab",
        "Data Structures Lab", "Introduction to Game Development Lab",
    ],
    "IT2C": [
        "Database Systems", "Operating System", "Data Structures & Algorithms",
        "Introduction to Game Development", "Ethics",
        "Fundamentals of Accounting for IT", "Environmental Science",
        "Database Systems Lab", "Operating System Lab",
        "Data Structures Lab", "Introduction to Game Development Lab",
    ],
    "IT3A": [
        "Computer Networks 1", "Platform Technologies", "Data Analytics",
        "Information and Project Management", "System Administration & Maintenance",
        "Fundamentals of Business Analytics", "Life and Works of Rizal",
        "Computer Networks 1 Lab", "Platform Technologies Lab",
        "Data Analytics Lab", "Information and Project Management Lab",
        "System Administration & Maintenance Lab", "Fundamentals of Business Analytics Lab",
    ],
    "IT3B": [
        "Computer Networks 1", "Platform Technologies", "Data Analytics",
        "Information and Project Management", "System Administration & Maintenance",
        "Fundamentals of Business Analytics", "Life and Works of Rizal",
        "Computer Networks 1 Lab", "Platform Technologies Lab",
        "Data Analytics Lab", "Information and Project Management Lab",
        "System Administration & Maintenance Lab", "Fundamentals of Business Analytics Lab",
    ],
    "IT3C": [
        "Computer Networks 1", "Platform Technologies", "Data Analytics",
        "Information and Project Management", "System Administration & Maintenance",
        "Fundamentals of Business Analytics", "Life and Works of Rizal",
        "Computer Networks 1 Lab", "Platform Technologies Lab",
        "Data Analytics Lab", "Information and Project Management Lab",
        "System Administration & Maintenance Lab", "Fundamentals of Business Analytics Lab",
    ],
    "IT4A": [
        "Social Issues & Ethics in Computing", "Information Assurance & Security 2",
        "Multimedia Systems", "IT Seminar", "Capstone Project Writing",
        "Information Assurance & Security 2 Lab", "Multimedia Systems Lab",
        "IT Seminar Lab", "Capstone Project Writing Lab",
    ],
    "IT4B": [
        "Social Issues & Ethics in Computing", "Information Assurance & Security 2",
        "Multimedia Systems", "IT Seminar", "Capstone Project Writing",
        "Information Assurance & Security 2 Lab", "Multimedia Systems Lab",
        "IT Seminar Lab", "Capstone Project Writing Lab",
    ],
    "IT4C": [
        "Social Issues & Ethics in Computing", "Information Assurance & Security 2",
        "Multimedia Systems", "IT Seminar", "Capstone Project Writing",
        "Information Assurance & Security 2 Lab", "Multimedia Systems Lab",
        "IT Seminar Lab", "Capstone Project Writing Lab",
    ],
}

# Lab sessions occupy 3 consecutive timeslots. The only requirement is that
# all 3 slots fit within the same day (p+2 <= 9 => p <= 7).
# Labs are allowed to span lunch slots.
valid_lab_starts = [1,4,7,10,13,16,19,22,25,28,31,34,37,40,43]

# Precompute subjects available per section to avoid repeated lookups.
lecture_subjects_in_section = {
    sec: [sub for sub in section_subjects.get(sec, []) if sub in lecture_subjects]
    for sec in sections
}
lab_subjects_in_section = {
    sec: [sub for sub in section_subjects.get(sec, []) if sub in lab_subjects]
    for sec in sections
}

# ----------------------------
# MODEL
# ----------------------------
model = LpProblem("Scheduling", LpMinimize)

# ----------------------------
# VARIABLES
# ----------------------------
valid_lecture_slots = [t for t in timeslots if t not in lunch_slots]

x = LpVariable.dicts("Lecture",
    [(r,t,sec,sub)
     for r in lecture_rooms
     for t in valid_lecture_slots
     for sec in sections
     for sub in lecture_subjects_in_section[sec]],
    cat="Binary"
)

z = LpVariable.dicts("Lab",
    [(r,t,sec,sub)
     for r in lab_rooms
     for t in valid_lab_starts
     for sec in sections
     for sub in lab_subjects_in_section[sec]],
    cat="Binary"
)

# ----------------------------
# INDEXING (FAST)
# ----------------------------
x_sec_t = defaultdict(list)
x_room_t = defaultdict(list)
x_sec_sub = defaultdict(list)
x_sec_sub_t = defaultdict(list)
x_room_sub = defaultdict(list)

for (r,t,sec,sub), var in x.items():
    x_sec_t[(sec,t)].append(var)
    x_room_t[(r,t)].append(var)
    x_sec_sub[(sec,sub)].append(var)
    x_sec_sub_t[(sec,sub,t)].append(var)
    x_room_sub[(sec,sub,r)].append(var)

z_sec_t = defaultdict(list)
z_room_t = defaultdict(list)
z_sec_sub = defaultdict(list)

for (r,t,sec,sub), var in z.items():
    for dt in range(3):
        z_sec_t[(sec,t+dt)].append(var)
        z_room_t[(r,t+dt)].append(var)
    z_sec_sub[(sec,sub)].append(var)

def occ(sec,t):
    return lpSum(x_sec_t[(sec,t)]) + lpSum(z_sec_t[(sec,t)])

# ----------------------------
# HARD CONSTRAINTS (FEASIBILITY FIRST)
# ----------------------------

# Room capacity
for r in lecture_rooms:
    for t in valid_lecture_slots:
        model += lpSum(x_room_t[(r,t)]) <= 1

for r in lab_rooms:
    for t in timeslots:
        model += lpSum(z_room_t[(r,t)]) <= 1

# Section constraint
for sec in sections:
    for t in timeslots:
        model += occ(sec,t) <= 1

# 🔥 STRICT UNIT ENFORCEMENT (DO NOT TOUCH)
for sec in sections:
    for sub in lecture_subjects_in_section[sec]:
        model += lpSum(x_sec_sub[(sec,sub)]) == lecture_subjects[sub]

    for sub in lab_subjects_in_section[sec]:
        model += lpSum(z_sec_sub[(sec,sub)]) == lab_subjects[sub]

# Lunch break
for sec in sections:
    for t in lunch_slots:
        model += occ(sec,t) == 0

# No same subject at same time across sections
for t in valid_lecture_slots:
    for sub in lecture_subjects:
        model += lpSum(
            x[(r,t,sec,sub)]
            for (r,t_,sec,sub_) in x
            if t_ == t and sub_ == sub
        ) <= 1

# ----------------------------
# OBJECTIVE (SAFE + FAST)
# ----------------------------

# 1. VACANT PENALTY (light)
vacant_penalty = lpSum(
    1 - occ(sec,t)
    for sec in sections
    for t in valid_lecture_slots
)

# 2. ROOM CONSISTENCY (VERY IMPORTANT)
room_penalty = []
for sec in sections:
    for sub in lecture_subjects_in_section[sec]:
        for r in lecture_rooms:
            used = LpVariable(f"room_used_{sec}_{sub}_{r}", cat="Binary")

            model += used >= lpSum(x_room_sub[(sec,sub,r)]) / 10

            room_penalty.append(used)

# 3. TIME CONSISTENCY (CORRECT LOGIC)
time_penalty = []

for sec in sections:
    for sub in lecture_subjects_in_section[sec]:
        for k in range(1,10):
            if k == 5:
                continue

            base = lpSum(x_sec_sub_t[(sec,sub,k)])  # Monday

            for d in range(1,5):
                t = d*9 + k
                curr = lpSum(x_sec_sub_t[(sec,sub,t)])

                p = LpVariable(f"time_{sec}_{sub}_{k}_{d}", lowBound=0)

                model += p >= base - curr
                model += p >= curr - base

                time_penalty.append(p)

# ----------------------------
# FINAL OBJECTIVE
# ----------------------------
model += (
    5 * vacant_penalty +
    50 * lpSum(room_penalty) +      # STRONG preference
    30 * lpSum(time_penalty)       # STRONG preference
)

# ----------------------------
# SOLVE
# ----------------------------
print("Solving...")
model.solve(PULP_CBC_CMD(msg=1, timeLimit=120, threads=4))

print("Status:", LpStatus[model.status])

# ----------------------------
# BUILD SCHEDULE
# ----------------------------
schedule = []

for (r,t,sec,sub) in x:
    if value(x[(r,t,sec,sub)]) == 1:
        schedule.append([sec, sub, r, t, "Lecture"])

for (r,t,sec,sub) in z:
    if value(z[(r,t,sec,sub)]) == 1:
        for dt in range(3):
            schedule.append([sec, sub, r, t+dt, "Lab"])

schedule_df = pd.DataFrame(
    schedule,
    columns=["Section","Subject","Room","Time","Type"]
)

# ----------------------------
# VERIFY UNITS (IMPORTANT)
# ----------------------------
print("\nUNIT CHECK:")
for sec in sections:
    for sub in section_subjects.get(sec, []):
        if sub in lecture_subjects:
            count = len(schedule_df[
                (schedule_df["Section"]==sec) &
                (schedule_df["Subject"]==sub)
            ])
            print(sec, sub, count, "/", lecture_subjects[sub])

        if sub in lab_subjects:
            count = len(schedule_df[
                (schedule_df["Section"]==sec) &
                (schedule_df["Subject"]==sub)
            ])
            print(sec, sub, count, "/", lab_subjects[sub]*3)



#Provides an excel sheet to visualize BLP result.
days = ["Mon","Tue","Wed","Thu","Fri"]
times = ["8-9","9-10","10-11","11-12","12-1","1-2","2-3","3-4","4-5"]

def decode(t):
    return days[(t-1)//9], times[(t-1)%9]

with pd.ExcelWriter("Full_Timetable.xlsx", engine="openpyxl") as writer:

    for sec in sections:
        grid = {time:{day:"" for day in days} for time in times}
        type_map = {}

        for _, row in schedule_df.iterrows():
            if row["Section"] == sec:
                d,t = decode(row["Time"])
                grid[t][d] = f"{row['Subject']} ({row['Room']})"
                type_map[(t,d)] = row["Type"]

        df_grid = pd.DataFrame(grid).T
        df_grid.to_excel(writer, sheet_name=sec)

        ws = writer.sheets[sec]

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 30

        lecture_fill = PatternFill(start_color="ADD8E6", fill_type="solid")
        lab_fill = PatternFill(start_color="90EE90", fill_type="solid")

        for i,time in enumerate(times, start=2):
            for j,day in enumerate(days, start=2):
                cell = ws.cell(row=i,column=j)
                if (time,day) in type_map:
                    cell.fill = lecture_fill if type_map[(time,day)]=="Lecture" else lab_fill

print("it is now saved at Full_Timetable.xlsx")